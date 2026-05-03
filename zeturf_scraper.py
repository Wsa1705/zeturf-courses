"""
zeturf_scraper.py  –  v5
========================================================
Structure Excel :
  - 1 onglet par réunion (hippodrome)
  - Colonnes : Date | N° Course | Hippodrome | Type | Classe/Groupe
               | N° Cheval | Cheval | Jockey | Entraîneur | Cote

N° Course = R{n} * 100 + C{m}  (ex: R1C4 → 104, R2C9 → 209)

Auteur : Claude / lkassist  |  2026-05-03
"""

import re
import time
import os
from datetime import datetime, timedelta

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ───────────────────────────────────────────────────────────────────
BASE_URL   = "https://www.zeturf.fr"
URL_PROG   = "https://www.zeturf.fr/fr/programmes-et-pronostics-du-jour"
TOMORROW   = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
DATE_FR    = (datetime.now() + timedelta(days=1)).strftime("%d-%b").replace(
                "Jan","jan").replace("Feb","fév").replace("Mar","mar").replace(
                "Apr","avr").replace("May","mai").replace("Jun","juin").replace(
                "Jul","juil").replace("Aug","aoû").replace("Sep","sep").replace(
                "Oct","oct").replace("Nov","nov").replace("Dec","déc")
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_XLS = os.path.join(OUTPUT_DIR, f"zeturf_{TOMORROW}.xlsx")

# ── OPTIONS DE FILTRAGE ───────────────────────────────────────────────────────
# True  = uniquement les hippodromes français
# False = TOUS les hippodromes du programme (13 le 04/05)
FRANCE_SEULEMENT = False

# Pays prioritaires pour filtrer les réunions françaises
PAYS_FR = ["vichy","chantilly","vincennes","auteuil","saint-cloud","deauville",
           "compiegne","rambouillet","lyon","bordeaux","caen","la-capelle",
           "meslay","maisons","paris-longchamp","longchamp","craon","clairefontaine",
           "pau","toulouse","strasbourg","angers","nantes","le-lion","pornichet"]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.zeturf.fr/fr",
}

session = requests.Session()
session.headers.update(HEADERS)

# ─── REGEX ────────────────────────────────────────────────────────────────────
RE_REUNION = re.compile(r"/reunion-du-jour/" + re.escape(TOMORROW) + r"/R(\d+)-([\w-]+)$")
RE_COURSE  = re.compile(r"/course-du-jour/\d{4}-\d{2}-\d{2}/R(\d+)C(\d+)-([\w-]+)$")


# ─── UTILITAIRE ───────────────────────────────────────────────────────────────
def get_page(url, label=""):
    print(f"  🌐 {label[:55]:55s}", end=" ", flush=True)
    try:
        r = session.get(url, timeout=15)
        if r.status_code != 200:
            print(f"❌ HTTP {r.status_code}")
            return None, None
        soup = BeautifulSoup(r.text, "html.parser")
        print(f"✅ {len(r.text):,} chars")
        return r, soup
    except Exception as e:
        print(f"❌ {e}")
        return None, None


def save_debug(soup, name):
    path = os.path.join(OUTPUT_DIR, name)
    with open(path, "w", encoding="utf-8") as f:
        f.write(soup.prettify())
    print(f"     💾 Debug → {path}")


# ─── P1 : RÉUNIONS J+1 ───────────────────────────────────────────────────────
def p1_reunions_demain():
    print(f"\n{'═'*60}")
    print(f"  P1 – Réunions du {TOMORROW}")
    print(f"{'═'*60}")

    _, soup = get_page(URL_PROG, "Programme du jour")
    if soup is None:
        return []

    reunions = []
    seen = set()

    for a in soup.find_all("a", href=True):
        href = a["href"]
        m = RE_REUNION.search(href)
        if not m or "#" in href:
            continue
        full  = href if href.startswith("http") else BASE_URL + href
        num_r = m.group(1)
        hippo = m.group(2).replace("-", " ").upper()
        if full not in seen:
            seen.add(full)
            reunions.append({
                "numero":     num_r,
                "hippodrome": hippo,
                "href":       full,
                "is_fr":      any(p in href.lower() for p in PAYS_FR),
            })

    # Trie : FR d'abord, puis par numéro
    reunions.sort(key=lambda x: (not x["is_fr"], int(x["numero"])))

    print(f"\n  {'R':4s} {'Hippodrome':30s} {'FR':4s}")
    print(f"  {'-'*4} {'-'*30} {'-'*4}")
    for r in reunions:
        flag = "🇫🇷" if r["is_fr"] else "🌍"
        print(f"  R{r['numero']:3s} {r['hippodrome']:30s} {flag}")

    return reunions


# ─── P2 : COURSES D'UNE RÉUNION ──────────────────────────────────────────────
def p2_courses(reunion):
    _, soup = get_page(reunion["href"], f"R{reunion['numero']} {reunion['hippodrome'][:30]}")
    if soup is None:
        return []

    courses = []
    seen = set()
    num_reunion = reunion["numero"]  # ex: "1", "2", "4"

    for a in soup.find_all("a", href=True):
        href = a["href"]
        m = RE_COURSE.search(href)
        if not m:
            continue
        num_r = m.group(1)
        num_c = m.group(2)
        slug  = m.group(3)

        # ✅ FILTRE CRITIQUE : ne garder QUE les courses de CETTE réunion
        if num_r != num_reunion:
            continue

        full = href if href.startswith("http") else BASE_URL + href
        num_course = int(num_r) * 100 + int(num_c)
        nom = slug.split("-", 2)[-1].replace("-", " ").title() if "-" in slug else slug

        if full not in seen:
            seen.add(full)
            courses.append({
                "num_r":      int(num_r),
                "num_c":      int(num_c),
                "num_course": num_course,
                "nom":        nom,
                "href":       full,
            })

    courses.sort(key=lambda x: x["num_c"])
    print(f"     → {len(courses)} course(s) pour R{num_reunion}")
    return courses


# ─── RÉFÉRENTIELS TYPE & CLASSE ───────────────────────────────────────────────

# Type de course : mots-clés → valeur normalisée (ordre important : plus précis d'abord)
TYPE_COURSE_MAP = [
    (r"\battel[eé]\b",   "Attelé"),
    (r"\bmont[eé]\b",    "Monté"),
    (r"\bhaies?\b",      "Haies"),
    (r"\bsteeple\b",     "Steeple"),
    (r"\bobstacle\b",    "Obstacle"),
    (r"\btrot\b",        "Trot"),
    (r"\bgalop\b",       "Galop"),
    (r"\bplat\b",        "Plat"),
]

# Liste exhaustive des Classe/Groupe PMU (ordre : plus long d'abord pour éviter
# qu'un sous-string matche avant la valeur complète)
CLASSES_PMU = [
    # Composés longs d'abord
    "Handicap / Classe 1", "Handicap / Classe 2", "Handicap / Classe 3",
    "Handicap / Classe 4", "Handicap / Classe 5", "Handicap / Classe 6",
    "Handicap / Groupe 3", "Handicap / Groupe 2", "Handicap / Groupe 1",
    "Handicap / Listed", "Handicap / Maiden", "Handicap / divisé",
    "Handicap / a", "Handicap / b", "Handicap / c", "Handicap / d",
    "Handicap / e", "Handicap / f", "Handicap / g",
    "Handicap de catégorie",
    "A condition / Classe 1", "A condition / Classe 2", "A condition / Classe 3",
    "A condition / Classe 4", "A condition / Listed",
    "A condition / Inédits", "A condition / Maiden", "A condition / Aqps",
    "A condition / a", "A condition / b", "A condition / c", "A condition / d",
    "A condition / e", "A condition / f",
    "A condition/ Classe 3",
    "A réclamer / Classe 3", "A réclamer / Classe 4",
    "A réclamer / Couse r", "A réclamer / r",
    "A réclamer / a", "A réclamer / b", "A réclamer / c", "A réclamer / d",
    "A réclamer / e", "A réclamer / f", "A réclamer / g",
    "Maiden / Classe 3", "Maiden / Classe 4", "Maiden / Classe 5",
    "Maiden / f", "Maiden / g",
    "Amateurs / g", "Apprentis / f", "Européenne / d",
    "Stakes / Groupe 1", "Stakes / Groupe 2", "Stakes / Groupe 3",
    "Clasico / Groupe 1",
    "Aqps / Classe 2",
    "Internationale / Groupe 1",
    "3 et 4 ans", "3 et 5 ans", "3 et 6 ans", "3 et 7 ans", "3 et 8 ans",
    "3 et 9 ans", "3 et 10 ans", "3 et 11 ans", "3 et 14 ans",
    "4 et 5 ans", "4 et 6 ans", "4 et 7 ans", "4 et 8 ans",
    "4 et 9 ans", "4 et 10 ans", "4 et 15 ans",
    "5 et 7 ans", "6 et 13 ans",
    # Simples
    "Groupe 1", "Groupe 2", "Groupe 3",
    "Classe 1", "Classe 2", "Classe 3", "Classe 4", "Classe 5",
    "Listed", "Maiden", "Handicap", "Stakes", "Clasico",
    "A condition", "A réclamer", "Nationale", "Internationale",
    "Amateurs", "Apprentis", "Aqps",
    "Inédits en obstacle", "Inédits",
    "3 ans et plus", "4 ans et plus", "5 ans et plus",
    "3 ans", "4 ans", "5 ans",
    "N/C",
    # Lettres seules (codes PMU)
    "a", "b", "c", "d", "e", "f", "g", "h", "r", "0",
]

# Pré-compile les regex pour chaque classe (insensible à la casse)
_CLASSES_RE = [
    (val, re.compile(r"(?<![/\w])" + re.escape(val) + r"(?![/\w])", re.IGNORECASE))
    for val in CLASSES_PMU
]


def extraire_type(texte):
    """Retourne le type de course normalisé depuis un texte brut."""
    t = texte.lower()
    for pattern, valeur in TYPE_COURSE_MAP:
        if re.search(pattern, t):
            return valeur
    return ""


def extraire_classe(texte):
    """
    Cherche dans 'texte' la première valeur de la liste CLASSES_PMU.
    Retourne la valeur canonique (casse PMU officielle).
    """
    for valeur, regex in _CLASSES_RE:
        if regex.search(texte):
            return valeur
    return ""


# ─── P3 : INFOS + PARTANTS D'UNE COURSE ─────────────────────────────────────
def p3_infos_partants(course, hippodrome):
    _, soup = get_page(
        course["href"],
        f"  C{course['num_c']:2d} {course['nom'][:35]}"
    )
    if soup is None:
        return {"type": "", "classe": "", "partants": []}

    type_course   = ""
    classe_course = ""

    # ── Stratégie 1 : balises dédiées aux métadonnées de la course
    META_SELECTORS = [
        "[class*='race-type']", "[class*='race-info']", "[class*='course-info']",
        "[class*='course-type']", "[class*='discipline']", "[class*='categorie']",
        "[class*='category']", "[class*='race-category']", "[class*='label']",
        "[class*='detail']", "[class*='info-course']", "[class*='fiche']",
        "dl dt", "dl dd", "table.info td", ".caracteristiques td",
    ]
    for sel in META_SELECTORS:
        for tag in soup.select(sel):
            txt = tag.get_text(separator=" ", strip=True)
            if not txt or len(txt) > 150:   # ignore les blocs trop grands
                continue
            if not type_course:
                type_course = extraire_type(txt)
            if not classe_course:
                classe_course = extraire_classe(txt)
            if type_course and classe_course:
                break
        if type_course and classe_course:
            break

    # ── Stratégie 2 : titre de la page (h1/h2/h3) souvent riche
    if not type_course or not classe_course:
        for tag in soup.find_all(["h1", "h2", "h3", "title"]):
            txt = tag.get_text(separator=" ", strip=True)
            if not type_course:
                type_course = extraire_type(txt)
            if not classe_course:
                classe_course = extraire_classe(txt)
            if type_course and classe_course:
                break

    # ── Stratégie 3 : scan de tous les textes courts (<= 80 chars)
    if not type_course or not classe_course:
        for tag in soup.find_all(["span", "div", "td", "li", "p"]):
            # Ignore les balises avec beaucoup d'enfants (blocs conteneurs)
            if len(tag.find_all()) > 3:
                continue
            txt = tag.get_text(separator=" ", strip=True)
            if not txt or len(txt) > 80:
                continue
            if not type_course:
                type_course = extraire_type(txt)
            if not classe_course:
                classe_course = extraire_classe(txt)
            if type_course and classe_course:
                break

    # ── Extraction : Partants
    partants = []
    seen_nums = set()

    # Méthode 1 : liens /N/infos
    for a in soup.find_all("a", href=True):
        href_a = a["href"]
        m = re.search(r"/(\d+)/infos$", href_a)
        if not m:
            continue
        num = m.group(1)
        if num in seen_nums:
            continue
        seen_nums.add(num)
        nom = a.get_text(strip=True)
        if not nom or nom == num:
            parent = a.find_parent(["td", "li", "div", "tr"])
            if parent:
                nom = parent.get_text(separator=" ", strip=True)[:40]
        partants.append({
            "numero":     num,
            "cheval":     nom,
            "jockey":     "",
            "entraineur": "",
            "cote":       "",
        })

    # Méthode 2 : tableau
    if len(partants) < 2:
        for sel in ["table tbody tr", "tr.cheval", "[class*='cheval']",
                    "[class*='horse']", "[class*='partant']"]:
            rows = soup.select(sel)
            if len(rows) >= 2:
                for row in rows:
                    cols = row.find_all(["td", "th"])
                    if len(cols) >= 2:
                        num = cols[0].get_text(strip=True)
                        if num and num not in seen_nums and num.isdigit():
                            seen_nums.add(num)
                            partants.append({
                                "numero":     num,
                                "cheval":     cols[1].get_text(strip=True) if len(cols) > 1 else "",
                                "jockey":     cols[2].get_text(strip=True) if len(cols) > 2 else "",
                                "entraineur": cols[3].get_text(strip=True) if len(cols) > 3 else "",
                                "cote":       cols[4].get_text(strip=True) if len(cols) > 4 else "",
                            })
                break

    partants.sort(key=lambda x: int(x["numero"]) if x["numero"].isdigit() else 99)

    nums = ", ".join(f"N°{p['numero']}" for p in partants)
    print(f"     → Type:{type_course or '?':10s} Classe:{classe_course[:25] or '?'} | {len(partants)} partants [{nums}]")

    return {"type": type_course, "classe": classe_course, "partants": partants}



# ─── ONGLET RÉSUMÉ ────────────────────────────────────────────────────────────
COLONNES_RESUME = [
    ("Date",           10),
    ("N° Course",       9),
    ("Hippodrome",     18),
    ("% Placés",        9),   # nb places PMU / nb partants
    ("Type",           10),
    ("Stat 2 %",        9),   # ← à définir / laisser vide
    ("Classe/Groupe",  22),
    ("Partants",        9),
    ("Places dispo",    9),   # nb chevaux placés (2 ou 3)
    ("Nom Course",     35),
]


def nb_places_pmu(nb_partants):
    """Retourne le nombre de chevaux placés selon les règles PMU."""
    if nb_partants <= 4:
        return 0
    elif nb_partants <= 7:
        return 2
    else:
        return 3

def pct_places(nb_partants):
    """% de partants placés (0–100 int), ou None si pas de pari placé."""
    n = nb_places_pmu(nb_partants)
    if nb_partants == 0:
        return None
    return round(n / nb_partants * 100)

def creer_onglet_resume(wb, toutes_reunions):
    """
    Crée un premier onglet "📋 RÉSUMÉ" avec une ligne par course.
    Colonnes : Date | N° Course | Hippodrome | Stat1% | Type | Stat2% | Classe | Partants | Heure | Nom
    Les colonnes Stat% sont vides — à renseigner selon ta méthodologie.
    """
    ws = wb.create_sheet(title="📋 RÉSUMÉ", index=0)  # premier onglet
    ws.freeze_panes = "A2"

    # Largeurs
    for col_idx, (_, width) in enumerate(COLONNES_RESUME, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 32

    # En-tête
    fill_hdr = PatternFill("solid", fgColor="1A1A2E")
    font_hdr = Font(bold=True, color="FFFFFF", size=10)
    for col_idx, (nom_col, _) in enumerate(COLONNES_RESUME, 1):
        c = ws.cell(row=1, column=col_idx, value=nom_col)
        c.fill = fill_hdr
        c.font = font_hdr
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Couleurs alternées par réunion
    PALETTE = [
        "EBF5FB", "E8F8F5", "FEF9E7", "F9EBEA",
        "F5EEF8", "EAFAF1", "FDF2E9", "EAF2FF",
        "EBF5FB", "F2F3F4", "FDFEFE", "E9F7EF",
        "FEF5E7",
    ]

    current_row = 2
    total_courses = 0

    for idx_r, item_r in enumerate(toutes_reunions):
        reunion = item_r["reunion"]
        courses = item_r["courses"]
        couleur_bg = PALETTE[idx_r % len(PALETTE)]
        fill_reu   = PatternFill("solid", fgColor=couleur_bg)

        # Couleur du N° de réunion dans le texte de hippodrome
        couleur_txt = COULEURS_REUNION[idx_r % len(COULEURS_REUNION)]
        font_hippo  = Font(bold=True, color=couleur_txt, size=9)

        for course_item in courses:
            course   = course_item["course"]
            type_c   = course_item["type"]
            classe_c = course_item["classe"]
            partants = course_item["partants"]
            nb_part  = len(partants)

            # Nom lisible court (20 premiers mots max)
            nom_course = course["nom"][:45]

            # Heure : non disponible dans le scraping actuel → vide
            heure = ""

            font_std = Font(size=9)
            font_num = Font(bold=True, size=9)

            def wr(col, val, font=font_std, align="left", fmt=None):
                c = ws.cell(row=current_row, column=col, value=val)
                c.fill = fill_reu
                c.font = font
                c.alignment = Alignment(horizontal=align, vertical="center")
                if fmt:
                    c.number_format = fmt

            pct   = pct_places(nb_part)
            n_pl  = nb_places_pmu(nb_part)
            # Formate le % avec signe
            pct_str = f"{pct}%" if pct is not None else "N/A"

            wr(1,  DATE_FR,               font_std,  "center")
            wr(2,  course["num_course"],   font_num,  "center")
            wr(3,  reunion["hippodrome"],  font_hippo,"left")
            wr(4,  pct_str,                font_num,  "center")   # % placés
            wr(5,  type_c,                 font_std,  "left")
            wr(6,  None,                   font_std,  "center")   # Stat2 → vide
            wr(7,  classe_c,               font_std,  "left")
            wr(8,  nb_part,                font_num,  "center")
            wr(9,  n_pl,                   font_num,  "center")   # places dispo
            wr(10, nom_course,             font_std,  "left")

            # Bordure légère
            thin = Side(style="thin", color="CCCCCC")
            for col in range(1, len(COLONNES_RESUME) + 1):
                ws.cell(row=current_row, column=col).border = Border(
                    bottom=thin, top=thin
                )

            ws.row_dimensions[current_row].height = 16
            current_row += 1
            total_courses += 1

        # Ligne séparatrice entre réunions (fond gris clair)
        fill_sep = PatternFill("solid", fgColor="D5D8DC")
        for col in range(1, len(COLONNES_RESUME) + 1):
            c = ws.cell(row=current_row, column=col)
            c.fill = fill_sep
        ws.row_dimensions[current_row].height = 4
        current_row += 1

    print(f"  📋 Onglet RÉSUMÉ : {total_courses} courses")
    return ws

# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────
# Couleurs par réunion (cycle)
COULEURS_REUNION = [
    "1F4E79",  # 1  bleu foncé    – VICHY
    "375623",  # 2  vert foncé    – COMPIEGNE
    "7B2C2C",  # 3  bordeaux      – MESLAY
    "4A235A",  # 4  violet        – CHANTILLY
    "7D6608",  # 5  ocre          – SANTIAGO
    "0B5345",  # 6  vert sapin    – PALERMO
    "154360",  # 7  bleu nuit     – ALICE SPRINGS
    "6E2F1A",  # 8  brun          – GUNNEDAH
    "1B4F72",  # 9  bleu marine   – FUNABASHI
    "4D5656",  # 10 ardoise       – VAAL
    "6C3483",  # 11 mauve         – NAGOYA
    "117A65",  # 12 vert forêt    – MANTORP
    "7E5109",  # 13 caramel       – WINDSOR
]
FILL_HEADER = "2E4057"   # En-tête colonnes
FILL_COURSE = "D9E1F2"  # Ligne de titre de course (bleu clair)


def style_header(ws, row, nb_cols):
    """Met en forme la ligne d'en-tête des colonnes."""
    fill = PatternFill("solid", fgColor=FILL_HEADER)
    font = Font(bold=True, color="FFFFFF", size=10)
    for col in range(1, nb_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_course_title(ws, row, nb_cols, color_hex):
    """Met en forme la ligne de titre d'une course."""
    fill = PatternFill("solid", fgColor=color_hex)
    font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        bottom=Side(style="medium", color="FFFFFF")
    )
    for col in range(1, nb_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border


def style_partant(ws, row, nb_cols, zebra=False):
    """Met en forme une ligne de partant."""
    fill = PatternFill("solid", fgColor="EBF5FB" if zebra else "FFFFFF")
    font = Font(size=9)
    for col in range(1, nb_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(vertical="center")


COLONNES = [
    ("Date",          10),
    ("N° Course",      9),
    ("Hippodrome",    18),
    ("Type",          10),
    ("Classe/Groupe", 22),
    ("N° Cheval",      9),
    ("Cheval",        24),
    ("Jockey",        20),
    ("Entraîneur",    20),
    ("Cote",           8),
]


def export_excel(toutes_reunions):
    """
    toutes_reunions : liste de dict
      { reunion: {...}, courses: [ {course, type, classe, partants} ] }
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_lignes = 0

    for idx_r, item_r in enumerate(toutes_reunions):
        reunion   = item_r["reunion"]
        courses   = item_r["courses"]
        couleur   = COULEURS_REUNION[idx_r % len(COULEURS_REUNION)]

        if not courses:
            continue

        # Nom de l'onglet
        hippo_court = reunion["hippodrome"].replace(" ", "-")[:25]
        nom_onglet  = re.sub(r"[\\/*?:\[\]]", "", f"R{reunion['numero']}-{hippo_court}")[:31]
        ws = wb.create_sheet(title=nom_onglet)
        ws.freeze_panes = "A2"

        # Largeurs de colonnes
        for col_idx, (_, width) in enumerate(COLONNES, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Hauteur de ligne standard
        ws.row_dimensions[1].height = 30

        # En-tête colonnes (ligne 1)
        for col_idx, (nom_col, _) in enumerate(COLONNES, 1):
            ws.cell(row=1, column=col_idx, value=nom_col)
        style_header(ws, 1, len(COLONNES))

        current_row = 2

        for course_item in courses:
            course   = course_item["course"]
            type_c   = course_item["type"]
            classe_c = course_item["classe"]
            partants = course_item["partants"]

            if not partants:
                continue

            # ── Ligne de titre de course (fusionnée)
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=len(COLONNES)
            )
            ws.cell(
                row=current_row, column=1,
                value=f"  C{course['num_c']}  –  {course['nom'].upper()}  "
                      f"(N° {course['num_course']})"
            )
            style_course_title(ws, current_row, len(COLONNES), couleur)
            ws.row_dimensions[current_row].height = 22
            current_row += 1

            # ── Lignes partants
            for z, p in enumerate(partants):
                ws.cell(row=current_row, column=1,  value=DATE_FR)
                ws.cell(row=current_row, column=2,  value=course["num_course"])
                ws.cell(row=current_row, column=3,  value=reunion["hippodrome"])
                ws.cell(row=current_row, column=4,  value=type_c)
                ws.cell(row=current_row, column=5,  value=classe_c)
                ws.cell(row=current_row, column=6,  value=p["numero"])
                ws.cell(row=current_row, column=7,  value=p["cheval"])
                ws.cell(row=current_row, column=8,  value=p["jockey"])
                ws.cell(row=current_row, column=9,  value=p["entraineur"])
                ws.cell(row=current_row, column=10, value=p["cote"])
                style_partant(ws, current_row, len(COLONNES), zebra=(z % 2 == 1))
                ws.row_dimensions[current_row].height = 16
                current_row += 1
                total_lignes += 1

            # Ligne vide entre les courses
            current_row += 1

    # ── Onglet résumé (inséré en premier)
    creer_onglet_resume(wb, toutes_reunions)

    # Gestion PermissionError si le fichier est ouvert dans Excel
    path_xls = OUTPUT_XLS
    for attempt in range(5):
        try:
            wb.save(path_xls)
            break
        except PermissionError:
            # Fichier ouvert → on génère un nom alternatif
            base, ext = os.path.splitext(OUTPUT_XLS)
            path_xls = f"{base}_v{attempt+2}{ext}"
            print(f"  ⚠️  Fichier verrouillé → essai avec {os.path.basename(path_xls)}")
    else:
        print("  ❌ Impossible de sauvegarder (fermez le fichier Excel ouvert)")
        return

    print(f"\n{'═'*60}")
    print(f"  ✅ Excel sauvegardé")
    print(f"     Fichier  : {path_xls}")
    print(f"     Onglets  : {len(wb.sheetnames)}")
    print(f"     Lignes   : {total_lignes} partants")
    print(f"{'═'*60}")


# ─── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"\n{'═'*60}")
    print(f"  ZEturf Scraper  v5  –  J+1 : {TOMORROW}")
    print(f"{'═'*60}")

    # P1 – Réunions de demain
    reunions = p1_reunions_demain()
    if not reunions:
        print("❌ Aucune réunion trouvée. Arrêt.")
        exit(1)

    # ── Sélection des réunions à scraper
    if FRANCE_SEULEMENT:
        reunions_cibles = [r for r in reunions if r["is_fr"]] or reunions
        print(f"\n  🇫🇷 Mode France uniquement")
    else:
        reunions_cibles = reunions  # TOUTES les réunions (FR + international)
        nb_fr  = sum(1 for r in reunions if r["is_fr"])
        nb_int = len(reunions) - nb_fr
        print(f"\n  🌍 Mode complet : {nb_fr} FR + {nb_int} international = {len(reunions_cibles)} réunion(s)")

    toutes_reunions = []

    for reunion in reunions_cibles:
        print(f"\n{'─'*60}")
        print(f"  R{reunion['numero']}  {reunion['hippodrome']}")
        print(f"{'─'*60}")

        # P2 – Courses de la réunion
        courses_meta = p2_courses(reunion)
        if not courses_meta:
            print("  ⚠️  Aucune course – réunion ignorée")
            continue

        time.sleep(0.8)

        # P3 – Infos + partants par course
        courses_data = []
        for course in courses_meta:
            infos = p3_infos_partants(course, reunion["hippodrome"])
            courses_data.append({
                "course":  course,
                "type":    infos["type"],
                "classe":  infos["classe"],
                "partants":infos["partants"],
            })
            time.sleep(0.6)

        toutes_reunions.append({
            "reunion": reunion,
            "courses": courses_data,
        })

    # Export Excel
    if toutes_reunions:
        export_excel(toutes_reunions)
    else:
        print("\n❌ Aucune donnée à exporter.")
