"""
zeturf_web.py  –  Interface web locale pour les courses ZEturf
==============================================================
Lancement : python zeturf_web.py
Accès     : http://localhost:5000

Installation : pip install flask openpyxl
"""

import os
import glob
import json
from datetime import datetime, timedelta
from flask import Flask, render_template_string, jsonify, request
import openpyxl

app = Flask(__name__)

# ─── CONFIG ───────────────────────────────────────────────────────────────────
DATA_DIR = os.path.dirname(os.path.abspath(__file__))

def find_latest_excel():
    """Trouve le fichier Excel le plus récent dans le dossier."""
    files = glob.glob(os.path.join(DATA_DIR, "zeturf_*.xlsx"))
    if not files:
        return None
    return max(files, key=os.path.getmtime)

def load_resume_from_excel(filepath):
    """Charge l'onglet RÉSUMÉ du fichier Excel."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    rows = []
    # Cherche l'onglet résumé (premier onglet ou celui qui contient RÉSUMÉ)
    sheet = None
    for name in wb.sheetnames:
        if "RÉSUMÉ" in name.upper() or "RESUME" in name.upper():
            sheet = wb[name]
            break
    if not sheet:
        sheet = wb[wb.sheetnames[0]]

    headers = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c else "" for c in row]
            continue
        if not any(row):  # ligne vide
            continue
        rows.append(dict(zip(headers, [str(c) if c is not None else "" for c in row])))
    wb.close()
    return headers, rows

# ─── TEMPLATE HTML ────────────────────────────────────────────────────────────
HTML = """<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ZEturf – Courses du {{ date }}</title>
<style>
  :root {
    --bg: #f8f9fa; --card: #ffffff; --border: #dee2e6;
    --text: #212529; --muted: #6c757d;
    --primary: #1F4E79; --accent: #0d6efd;
    --fr: #E8F5E9; --int: #E3F2FD;
    --row-hover: #f0f4ff;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: system-ui, sans-serif; background: var(--bg); color: var(--text); font-size: 14px; }

  header { background: var(--primary); color: #fff; padding: 14px 24px;
    display: flex; align-items: center; gap: 16px; }
  header h1 { font-size: 18px; font-weight: 600; }
  header .sub { font-size: 12px; opacity: .75; }
  .badge { background: rgba(255,255,255,.2); border-radius: 20px;
    padding: 2px 10px; font-size: 12px; }

  .toolbar { display: flex; flex-wrap: wrap; gap: 10px; padding: 14px 24px;
    background: var(--card); border-bottom: 1px solid var(--border);
    align-items: center; }
  .toolbar input, .toolbar select {
    border: 1px solid var(--border); border-radius: 6px;
    padding: 6px 10px; font-size: 13px; background: var(--bg); }
  .toolbar input { width: 200px; }
  .count { margin-left: auto; font-size: 12px; color: var(--muted); }

  .table-wrap { overflow-x: auto; padding: 0 24px 24px; margin-top: 16px; }
  table { border-collapse: collapse; width: 100%; background: var(--card);
    border-radius: 10px; overflow: hidden; box-shadow: 0 1px 4px rgba(0,0,0,.07); }
  thead tr { background: var(--primary); color: #fff; }
  th { padding: 10px 12px; text-align: left; font-weight: 600;
    font-size: 12px; white-space: nowrap; cursor: pointer; user-select: none; }
  th:hover { background: rgba(255,255,255,.15); }
  th .sort-icon { margin-left: 4px; opacity: .5; }
  th.sorted .sort-icon { opacity: 1; }

  td { padding: 8px 12px; border-bottom: 1px solid var(--border); white-space: nowrap; }
  tr:last-child td { border-bottom: none; }
  tr:hover td { background: var(--row-hover); }

  .tag { display: inline-block; border-radius: 4px; padding: 1px 8px;
    font-size: 11px; font-weight: 600; }
  .tag-plat     { background: #E3F2FD; color: #0d47a1; }
  .tag-attele   { background: #E8F5E9; color: #1b5e20; }
  .tag-haies    { background: #FFF8E1; color: #f57f17; }
  .tag-obstacle { background: #FCE4EC; color: #880e4f; }
  .tag-steeple  { background: #F3E5F5; color: #4a148c; }
  .tag-monte    { background: #E0F7FA; color: #006064; }
  .tag-trot     { background: #FFF3E0; color: #e65100; }
  .tag-default  { background: #F5F5F5; color: #424242; }

  .pct { font-weight: 700; font-size: 13px; }
  .pct-high { color: #2e7d32; }
  .pct-mid  { color: #f57f17; }
  .pct-low  { color: #c62828; }

  .num-course { font-weight: 700; color: var(--accent); font-size: 13px; }
  .hippo { font-weight: 600; }

  .no-data { text-align: center; padding: 60px; color: var(--muted); }
  .file-info { font-size: 11px; color: var(--muted); padding: 8px 24px; }

  .refresh-btn { background: var(--accent); color: #fff; border: none;
    border-radius: 6px; padding: 6px 14px; cursor: pointer; font-size: 13px; }
  .refresh-btn:hover { opacity: .85; }

  @media (max-width: 640px) {
    header { padding: 10px 14px; }
    .toolbar { padding: 10px 14px; }
    .table-wrap { padding: 0 10px 20px; }
  }
</style>
</head>
<body>

<header>
  <div>
    <h1>🏇 ZEturf – Programme J+1</h1>
    <div class="sub">Courses du {{ date }}</div>
  </div>
  <span class="badge">{{ nb_courses }} courses</span>
  <span class="badge">{{ nb_reunions }} réunions</span>
</header>

<div class="toolbar">
  <input type="text" id="search" placeholder="🔍 Rechercher…" oninput="filterTable()">
  <select id="filterType" onchange="filterTable()">
    <option value="">Tous les types</option>
    <option>Plat</option><option>Attelé</option><option>Trot</option>
    <option>Haies</option><option>Obstacle</option><option>Steeple</option><option>Monté</option>
  </select>
  <select id="filterHippo" onchange="filterTable()">
    <option value="">Tous les hippodromes</option>
    {% for h in hippodromes %}<option>{{ h }}</option>{% endfor %}
  </select>
  <select id="filterClasse" onchange="filterTable()">
    <option value="">Toutes les classes</option>
    {% for c in classes %}<option>{{ c }}</option>{% endfor %}
  </select>
  <button class="refresh-btn" onclick="location.reload()">↻ Actualiser</button>
  <span class="count" id="count-label"></span>
</div>

{% if file_info %}
<div class="file-info">📁 {{ file_info }}</div>
{% endif %}

<div class="table-wrap">
{% if rows %}
<table id="mainTable">
  <thead>
    <tr>
      <th onclick="sortTable(0)">Date <span class="sort-icon">↕</span></th>
      <th onclick="sortTable(1)" class="sorted">N° Course <span class="sort-icon">↑</span></th>
      <th onclick="sortTable(2)">Hippodrome <span class="sort-icon">↕</span></th>
      <th onclick="sortTable(3)">% Placés <span class="sort-icon">↕</span></th>
      <th onclick="sortTable(4)">Type <span class="sort-icon">↕</span></th>
      <th>Stat 2%</th>
      <th onclick="sortTable(6)">Classe/Groupe <span class="sort-icon">↕</span></th>
      <th onclick="sortTable(7)">Partants <span class="sort-icon">↕</span></th>
      <th>Places</th>
      <th>Nom de la course</th>
    </tr>
  </thead>
  <tbody id="tableBody">
  {% for r in rows %}
  <tr
    data-type="{{ r.get('Type','') | lower }}"
    data-hippo="{{ r.get('Hippodrome','') | lower }}"
    data-classe="{{ r.get('Classe/Groupe','') | lower }}"
    data-search="{{ (r.values() | join(' ')) | lower }}">
    <td>{{ r.get('Date','') }}</td>
    <td><span class="num-course">{{ r.get('N° Course','') }}</span></td>
    <td><span class="hippo">{{ r.get('Hippodrome','') }}</span></td>
    <td>
      {% set pct = r.get('% Placés','') %}
      {% if pct and pct != 'N/A' %}
        {% set pval = pct | replace('%','') | int(0) %}
        <span class="pct {% if pval >= 35 %}pct-high{% elif pval >= 25 %}pct-mid{% else %}pct-low{% endif %}">
          {{ pct }}
        </span>
      {% else %}<span class="pct" style="color:var(--muted)">N/A</span>{% endif %}
    </td>
    <td>
      {% set t = r.get('Type','') | lower %}
      <span class="tag tag-{% if t %}{{ t | replace('é','e') }}{% else %}default{% endif %}">
        {{ r.get('Type','') or '—' }}
      </span>
    </td>
    <td>{{ r.get('Stat 2 %','') or '' }}</td>
    <td>{{ r.get('Classe/Groupe','') or '—' }}</td>
    <td style="text-align:center;font-weight:700">{{ r.get('Partants','') }}</td>
    <td style="text-align:center">{{ r.get('Places dispo','') }}</td>
    <td style="color:var(--muted);font-size:12px;max-width:280px;overflow:hidden;text-overflow:ellipsis">
      {{ r.get('Nom Course','') }}
    </td>
  </tr>
  {% endfor %}
  </tbody>
</table>
{% else %}
<div class="no-data">
  <p>Aucune donnée disponible.</p>
  <p style="margin-top:8px;font-size:12px">Lance <code>zeturf_scraper.py</code> pour générer le fichier Excel.</p>
</div>
{% endif %}
</div>

<script>
const tbody = document.getElementById('tableBody');
let sortDir = [1,1,1,1,1,1,1,1];

function filterTable() {
  const q     = document.getElementById('search').value.toLowerCase();
  const type  = document.getElementById('filterType').value.toLowerCase();
  const hippo = document.getElementById('filterHippo').value.toLowerCase();
  const cl    = document.getElementById('filterClasse').value.toLowerCase();
  let count = 0;
  tbody.querySelectorAll('tr').forEach(tr => {
    const search = tr.dataset.search || '';
    const ok =
      (!q     || search.includes(q)) &&
      (!type  || tr.dataset.type  === type) &&
      (!hippo || tr.dataset.hippo === hippo) &&
      (!cl    || tr.dataset.classe.includes(cl));
    tr.style.display = ok ? '' : 'none';
    if (ok) count++;
  });
  const lbl = document.getElementById('count-label');
  if (lbl) lbl.textContent = count + ' course(s) affichée(s)';
}

function sortTable(col) {
  const rows = Array.from(tbody.querySelectorAll('tr'));
  const dir  = sortDir[col];
  rows.sort((a, b) => {
    const av = a.cells[col]?.textContent.trim() || '';
    const bv = b.cells[col]?.textContent.trim() || '';
    const an = parseFloat(av), bn = parseFloat(bv);
    if (!isNaN(an) && !isNaN(bn)) return (an - bn) * dir;
    return av.localeCompare(bv, 'fr') * dir;
  });
  sortDir[col] *= -1;
  document.querySelectorAll('th').forEach(th => th.classList.remove('sorted'));
  document.querySelectorAll('th')[col].classList.add('sorted');
  rows.forEach(r => tbody.appendChild(r));
  filterTable();
}

// Comptage initial
window.addEventListener('load', () => {
  filterTable();
  sortTable(1); sortTable(1); // tri par N° Course asc
});
</script>
</body>
</html>
"""

# ─── ROUTES ───────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    filepath = find_latest_excel()
    rows, headers, hippodromes, classes, file_info = [], [], [], [], ""
    date_str = (datetime.now() + timedelta(days=1)).strftime("%d/%m/%Y")
    nb_courses = nb_reunions = 0

    if filepath:
        file_info = f"{os.path.basename(filepath)} — modifié le {datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%d/%m/%Y %H:%M')}"
        try:
            headers, rows = load_resume_from_excel(filepath)
            hippodromes = sorted(set(r.get("Hippodrome","") for r in rows if r.get("Hippodrome","")))
            classes     = sorted(set(r.get("Classe/Groupe","") for r in rows if r.get("Classe/Groupe","")))
            nb_courses  = len(rows)
            nb_reunions = len(hippodromes)
        except Exception as e:
            file_info += f" ⚠️ Erreur lecture : {e}"
    else:
        file_info = "Aucun fichier zeturf_*.xlsx trouvé dans " + DATA_DIR

    return render_template_string(
        HTML,
        rows=rows, headers=headers,
        date=date_str,
        nb_courses=nb_courses, nb_reunions=nb_reunions,
        hippodromes=hippodromes, classes=classes,
        file_info=file_info,
    )

@app.route("/api/data")
def api_data():
    """Endpoint JSON pour intégration future."""
    filepath = find_latest_excel()
    if not filepath:
        return jsonify({"error": "Aucun fichier trouvé"}), 404
    _, rows = load_resume_from_excel(filepath)
    return jsonify(rows)

# ─── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 55)
    print("  🏇 ZEturf Web  –  http://localhost:5000")
    print(f"  📁 Dossier : {DATA_DIR}")
    latest = find_latest_excel()
    if latest:
        print(f"  📊 Fichier  : {os.path.basename(latest)}")
    else:
        print("  ⚠️  Aucun fichier Excel trouvé – lance d'abord le scraper")
    print("=" * 55)
    app.run(debug=False, host="127.0.0.1", port=5000)
